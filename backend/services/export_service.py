"""
services/export_service.py — Export Excel des tickets et métriques
"""
from __future__ import annotations

import io
from datetime import date, datetime
from typing import Any

from core.config import Settings


def _parse_date(val: Any) -> date | None:
    if not val or val in ("NULL", "0000-00-00 00:00:00"):
        return None
    try:
        return datetime.fromisoformat(str(val).replace(" ", "T")).date()
    except ValueError:
        return None


class ExportService:
    def __init__(self, settings: Settings):
        self._s = settings

    def export_tickets_excel(self, tickets: list[dict]) -> bytes:
        """Génère un fichier Excel avec tous les tickets."""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            raise RuntimeError("openpyxl non installé — ajoutez-le dans requirements.txt")

        wb = openpyxl.Workbook()

        # ── Feuille 1 : Tickets ──────────────────────────────────
        ws = wb.active
        ws.title = "Tickets"

        header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)

        headers = [
            "ID", "Titre", "Statut", "Priorité",
            "Date création", "Date résolution", "Délai (jours)",
            "En retard", "Acheteur", "Projet", "Catégorie ID",
        ]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        for row, t in enumerate(tickets, 2):
            created = _parse_date(t.get("date"))
            closed = _parse_date(t.get("closedate")) or _parse_date(t.get("solvedate"))
            delay = (closed - created).days if created and closed else None
            status = t.get("status", 0)
            is_late = (
                status not in self._s.resolved_statuses
                and _parse_date(t.get("time_to_resolve")) is not None
                and date.today() > _parse_date(t.get("time_to_resolve"))
            )

            ws.cell(row=row, column=1, value=t.get("id"))
            ws.cell(row=row, column=2, value=t.get("name", ""))
            ws.cell(row=row, column=3, value=self._s.status_map.get(status, "Inconnu"))
            ws.cell(row=row, column=4, value=t.get("priority", 0))
            ws.cell(row=row, column=5, value=str(created) if created else "")
            ws.cell(row=row, column=6, value=str(closed) if closed else "")
            ws.cell(row=row, column=7, value=delay)
            ws.cell(row=row, column=8, value="Oui" if is_late else "Non")
            ws.cell(row=row, column=9, value=t.get("_buyer_name", "Non assigné"))
            ws.cell(row=row, column=10, value=t.get("_project_name", "Sans projet"))
            ws.cell(row=row, column=11, value=t.get("itilcategories_id", 0))

        # Largeur colonnes
        col_widths = [8, 50, 20, 12, 16, 16, 14, 12, 30, 30, 14]
        for col, width in enumerate(col_widths, 1):
            ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = width

        # ── Feuille 2 : Résumé ───────────────────────────────────
        ws2 = wb.create_sheet("Résumé")
        ws2.append(["Métrique", "Valeur"])
        total = len(tickets)
        resolved = sum(1 for t in tickets if t.get("status") in self._s.resolved_statuses)
        ws2.append(["Total tickets", total])
        ws2.append(["Résolus", resolved])
        ws2.append(["Ouverts", total - resolved])
        ws2.append(["Taux réalisation (%)", round(resolved / total * 100, 1) if total else 0])
        ws2.append(["Urgents", sum(1 for t in tickets if t.get("priority", 0) in self._s.urgent_priorities)])
        ws2.append(["Date export", str(date.today())])

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.read()